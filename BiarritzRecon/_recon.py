import openpyxl, re, os, sys
from datetime import datetime
from collections import defaultdict

BASE = r'C:\GitHubWS\CareerDNA\BiarritzRecon'
MONTHS = ['April2025','May2025','June2025','July2025','Aug2025','Sept2025',
          'Oct2025','Nov2025','Dec2025','Jan2026','Feb2026','Mar2026']

def dmy(s):
    for fmt in ('%d/%m/%y','%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%d/%m/%Y %H:%M:%S'):
        try: return datetime.strptime(str(s).strip(), fmt)
        except Exception: pass
    return None

def num(v):
    try: return float(v)
    except Exception: return None

# ---------- 1. Parse bank statements ----------
def parse_statement(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    txns = []
    started = False
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        if isinstance(c0, str) and set(c0.strip()) == {'*'} and len(c0.strip()) > 8:
            started = True
            continue
        if not started: continue
        if not isinstance(c0, str) or not re.match(r'^\d{2}/\d{2}/\d{2,4}$', c0.strip()):
            continue
        date = dmy(c0)
        narr = row[1] if len(row) > 1 else None
        ref  = row[2] if len(row) > 2 else None
        wd   = row[4] if len(row) > 4 else None
        dp   = row[5] if len(row) > 5 else None
        bal  = row[6] if len(row) > 6 else None
        txns.append({'date': date, 'narr': str(narr or ''), 'ref': str(ref or '').strip(),
                     'wd': float(wd) if isinstance(wd,(int,float)) else 0.0,
                     'dp': float(dp) if isinstance(dp,(int,float)) else 0.0,
                     'bal': float(bal) if isinstance(bal,(int,float)) else None})
    return txns

bank = []
per_month = {}
for m in MONTHS:
    t = parse_statement(os.path.join(BASE, m + '.xlsx'))
    per_month[m] = t
    for x in t: x['month'] = m
    bank.extend(t)

credits = [x for x in bank if x['dp'] > 0]
debits  = [x for x in bank if x['wd'] > 0]

print('='*70)
print('BANK STATEMENTS PARSED')
for m in MONTHS:
    t = per_month[m]
    tot_dp = sum(x['dp'] for x in t); tot_wd = sum(x['wd'] for x in t)
    print(f'  {m:<10} txns={len(t):>3}  credits=Rs {tot_dp:>12,.2f}  debits=Rs {tot_wd:>12,.2f}')
tot_credits, tot_debits = sum(x['dp'] for x in credits), sum(x['wd'] for x in debits)
print(f'  TOTAL      credits=Rs {tot_credits:,.2f}   debits=Rs {tot_debits:,.2f}')

print('\nBALANCE CONTINUITY CHECK')
prev_close = None; ok = True
for m in MONTHS:
    t = per_month[m]
    if not t: continue
    bals = [x['bal'] for x in t if x['bal'] is not None]
    close = bals[-1] if bals else None
    imp_open = round(close - sum(x['dp'] for x in t) + sum(x['wd'] for x in t), 2) if close is not None else None
    flag = ''
    if prev_close is not None and imp_open is not None and abs(imp_open - prev_close) > 0.01:
        flag = f'  <-- GAP vs prior close Rs {prev_close:,.2f}'; ok = False
    print(f'  {m:<10} implied opening=Rs {imp_open!s:>12}  closing=Rs {close!s:>12}{flag}')
    prev_close = close
print('  Continuity:', 'OK' if ok else 'BREAKS FOUND')

# ---------- 2. Parse dues collection ----------
wb = openpyxl.load_workbook(os.path.join(BASE,'Dues Collection Report_25Aug26.xlsx'), data_only=True)
ws = wb.active
dues = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None or not str(row[0]).strip().isdigit(): continue
    dues.append({'sl':int(row[0]), 'date':row[1], 'house':row[2], 'rtype':row[3],
                 'receipt':str(row[5]), 'payee':row[7], 'ptype':str(row[8] or '').lower(),
                 'amount':num(row[9]) or 0.0, 'bank_date':row[10],
                 'refno':str(row[11]).strip() if row[11] else '',
                 'ledger':str(row[12] or ''), 'status':str(row[13] or '')})
print('\n' + '='*70)
print(f'DUES COLLECTION: {len(dues)} receipts, total Rs {sum(d["amount"] for d in dues):,.2f}')

# ---------- 3. Match dues -> bank credits by UTR/ref ----------
matched_pairs, unmatched_dues = [], []
used_bank = set()
by_ref_dues = defaultdict(list)
for d in dues:
    if d['refno']:
        by_ref_dues[d['refno']].append(d)
    else:
        unmatched_dues.append((d, 'no UTR/reference (cash or manual)'))

for ref in sorted(by_ref_dues):
    ds = by_ref_dues[ref]
    hit = next(((i,x) for i,x in enumerate(credits) if x['ref'] == ref and i not in used_bank), None)
    if hit is not None:
        i, bx = hit
        used_bank.add(i)
        due_sum = sum(d['amount'] for d in ds)
        matched_pairs.append((ds, bx))
        if abs(due_sum - bx['dp']) >= 0.01:
            for d in ds: unmatched_dues.append((d, f'AMOUNT MISMATCH: receipts sum Rs {due_sum:,.2f} vs bank Rs {bx["dp"]:,.2f}'))
    else:
        for d in ds: unmatched_dues.append((d, f'UTR {ref} NOT FOUND in bank statements'))

unmatched_bank_credits = [x for i,x in enumerate(credits) if i not in used_bank]
cash_total = sum(d['amount'] for d in dues if 'cash' in d['ledger'].lower() or 'cash' in d['ptype'])
matched_due_amt = sum(sum(d['amount'] for d in ds) for ds,_ in matched_pairs)

def classify_credit(x):
    n = x['narr'].upper()
    if 'AUTO_REDEEM' in n or 'PRIN AND INT' in n: return 'FD auto-renewal (principal+interest re-deposit)'
    if 'INTEREST CREDIT' in n: return 'FD interest credit'
    if 'UPI SETTLEMENT' in n: return 'UPI settlement (petty receipt)'
    if 'CHQ ' in n or 'CTS' in n: return 'Cheque/CTS deposit'
    if 'NEFT' in n: return 'NEFT credit'
    if 'IMPS' in n: return 'IMPS credit'
    return 'Other'

credit_classes = defaultdict(lambda: [0, 0.0])
for x in unmatched_bank_credits:
    c = classify_credit(x)
    credit_classes[c][0] += 1
    credit_classes[c][1] += x['dp']

print('\nDUES <-> BANK CREDIT MATCHING')
print(f'  Matched groups : {len(matched_pairs)}  covering receipts worth Rs {matched_due_amt:,.2f}')
print(f'  Cash receipts  : Rs {cash_total:,.2f} (expected - never hits this bank a/c)')
print(f'  Unmatched dues : {len(unmatched_dues)} items worth Rs {sum(d["amount"] for d,_ in unmatched_dues):,.2f}')

if unmatched_dues:
    print('\n  UNMATCHED RECEIPTS:')
    for d,r in sorted(unmatched_dues, key=lambda t:str(t[1])):
        print(f'    Rcpt#{d["receipt"]:<6} {str(d["date"])[:10]} {str(d["house"]):<7} Rs {d["amount"]:>10,.2f}  [{r}]')

if unmatched_bank_credits:
    print(f'\n  BANK CREDITS WITHOUT ANY RECEIPT ({len(unmatched_bank_credits)}, Rs {sum(x["dp"] for x in unmatched_bank_credits):,.2f}):')
    for c,(cnt,amt) in sorted(credit_classes.items(), key=lambda kv:-kv[1][1]):
        print(f'    {c:<50} {cnt:>3} items  Rs {amt:,.2f}')
    print('\n  Detail (non-FD items):')
    for x in unmatched_bank_credits:
        if classify_credit(x).startswith('FD auto'): continue
        print(f'    {x["date"].strftime("%d/%m/%y")} Rs {x["dp"]:>10,.2f}  [{classify_credit(x)[:20]}] {x["narr"][:80]}')



# ---------- 4. Parse expense report & match to bank debits ----------
wb = openpyxl.load_workbook(os.path.join(BASE,'Expense-Report_details_25Aug26.xlsx'), data_only=True)
ws = wb.active
exp = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None or not str(row[0]).strip().isdigit(): continue
    exp.append({'vno':str(row[0]), 'date':dmy(row[2]), 'vendor':row[3], 'ac':row[6],
                'net_inv':num(row[18]) or 0.0, 'net_pay':num(row[20]) or 0.0,
                'settled':num(row[21]) or 0.0, 'bal':num(row[22]) or 0.0,
                'st_fy':str(row[23] or ''), 'st_aug':str(row[24] or '')})
print('\n' + '='*70)
print(f'EXPENSE REPORT: {len(exp)} bills, net payable Rs {sum(e["net_pay"] for e in exp):,.2f}, '
      f'settled Rs {sum(e["settled"] for e in exp):,.2f}, unpaid balance Rs {sum(e["bal"] for e in exp):,.2f}')

deb_avail = list(range(len(debits)))
unmatched_exp = []
matched_exp_amt = 0.0
for e in sorted(exp, key=lambda e:-e['settled']):
    if e['settled'] <= 0 or e['date'] is None: continue
    win = [i for i in deb_avail if abs((debits[i]['date'] - e['date']).days) <= 45
           and abs(debits[i]['wd'] - e['settled']) < 0.01]
    win.sort(key=lambda i: abs((debits[i]['date'] - e['date']).days))
    if len(win) >= 1:
        deb_avail.remove(win[0])
        matched_exp_amt += e['settled']
    else:
        unmatched_exp.append(e)

print('\nEXPENSES <-> BANK DEBIT MATCHING (exact amount, +/-45 days)')
print(f'  Matched   : Rs {matched_exp_amt:,.2f}')
print(f'  Unmatched : {len(unmatched_exp)} payments worth Rs {sum(e["settled"] for e in unmatched_exp):,.2f}')
print(f'  Bank debits with no bill match: {len(deb_avail)} items worth Rs {sum(debits[i]["wd"] for i in deb_avail):,.2f}')

if unmatched_exp:
    print('\n  SETTLED PAYMENTS NOT TRACED TO A BANK DEBIT:')
    for e in unmatched_exp[:40]:
        print(f'    V#{e["vno"]:<6} {str(e["date"])[:10]} {str(e["vendor"])[:40]:<40} Rs {e["settled"]:>10,.2f}')

print('\n  BANK DEBITS WITH NO MATCHING BILL (sample):')
for i in deb_avail[:40]:
    x = debits[i]
    print(f'    {x["date"].strftime("%d/%m/%y")} Rs {x["wd"]:>10,.2f}  {x["narr"][:80]}')

print('\n' + '='*70)
print('SUMMARY')
print(f'  Total bank credits : Rs {tot_credits:,.2f}')
print(f'  Total bank debits  : Rs {tot_debits:,.2f}')
print(f'  Net movement       : Rs {tot_credits-tot_debits:,.2f}')

# ---------- 5. Secondary match: unmatched receipts by amount+date ----------
print('\n' + '='*70)
print('SECONDARY TRACE OF UNMATCHED RECEIPTS (by exact amount, +/-15 days)')
leftover_bank = unmatched_bank_credits[:]
still_missing = []
traced_amt = 0.0
for d, r in list(unmatched_dues):
    hit = next((x for x in leftover_bank
                if abs(x['dp'] - d['amount']) < 0.01 and x['date'] and str(d['date'])[:10]
                and abs((x['date'] - datetime.strptime(str(d['date'])[:10], '%Y-%m-%d')).days) <= 15), None)
    if hit:
        leftover_bank.remove(hit)
        traced_amt += d['amount']
        print(f'  TRACED   Rcpt#{d["receipt"]:<6} {str(d["house"]):<7} Rs {d["amount"]:>10,.2f} -> bank {hit["date"].strftime("%d/%m/%y")} {hit["narr"][:60]}')
    else:
        still_missing.append((d, r))
print(f'  Traced by amount/date : Rs {traced_amt:,.2f}')
print(f'  GENUINELY UNTRACED    : {len(still_missing)} receipts worth Rs {sum(d["amount"] for d,_ in still_missing):,.2f}')
for d,r in still_missing:
    print(f'    Rcpt#{d["receipt"]:<6} {str(d["date"])[:10]} {str(d["house"]):<7} Rs {d["amount"]:>10,.2f}  [{r}]')

# ---------- 6. Classify unexplained bank debits ----------
def classify_debit(x):
    n = x['narr'].upper()
    if 'FD THROUGH NET' in n or 'AUTO_REDEEM' in n or 'PRIN AND INT' in n or 'LIEN' in n: return 'Fixed Deposits (placements/renewals)'
    if 'SBIEPY' in n or 'MSEB' in n or 'ELECT' in n: return 'Electricity bills (MSEB)'
    if 'PMC' in n or 'MUNICIPAL' in n or 'PROPERTY TAX' in n or 'BILLDKPUNEMUNICIPALC' in n: return 'Municipal / property taxes'
    if 'CBDT' in n or 'INCOME TAX' in n or 'TDS' in n or 'ITDTAX' in n: return 'Income tax / TDS'
    if 'CHQ DEP RET' in n or 'I/W CHQ RET' in n or 'CHQ RET' in n: return 'Inward cheque RETURNS (bounced)'
    if 'CHQ PAID' in n or 'CHQ ' in n: return 'Cheque payments (vendors)'
    if 'UPI' in n: return 'UPI payments'
    if 'FT - DR' in n or 'NEFT' in n or 'IMPS' in n: return 'Fund transfers (NEFT/FT/IMPS)'
    if 'SELF' in n: return 'Self / cash withdrawals'
    return 'Other'

debit_classes = defaultdict(lambda: [0, 0.0])
for i in deb_avail:
    c = classify_debit(debits[i])
    debit_classes[c][0] += 1
    debit_classes[c][1] += debits[i]['wd']
print('\n' + '='*70)
print('BANK DEBITS WITH NO VENDOR BILL, CLASSIFIED')
for c,(cnt,amt) in sorted(debit_classes.items(), key=lambda kv:-kv[1][1]):
    print(f'  {c:<42} {cnt:>3} items  Rs {amt:>13,.2f}')

# recon status tally from dues report
st = defaultdict(lambda: [0,0.0])
for d in dues:
    key = d['status'].split(' (')[0]
    st[key][0]+=1; st[key][1]+=d['amount']
print('\nDUES REPORT RECONCILIATION STATUS (as recorded by society software)')
for k,(cnt,amt) in sorted(st.items(), key=lambda kv:-kv[1][1]):
    print(f'  {k:<40} {cnt:>3} receipts  Rs {amt:>13,.2f}')

