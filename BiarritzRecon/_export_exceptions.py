"""Exports reconciliation exception lists to Excel for follow-up."""
import io, sys, os, contextlib
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = r'C:\GitHubWS\CareerDNA\BiarritzRecon'
sys.path.insert(0, BASE)

# Run the recon engine silently and reuse its parsed/matched data
_buf = io.StringIO()
with contextlib.redirect_stdout(_buf):
    import _recon as R

MONEY = '#,##0.00'
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF')
RED_FONT = Font(color='C00000', bold=True)

def make_sheet(wb, title, headers, rows, widths, money_cols=(), tab_color=None):
    ws = wb.create_sheet(title[:31])
    if tab_color: ws.sheet_properties.tabColor = tab_color
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical='center')
    for row in rows:
        ws.append(list(row))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in money_cols:
        for r in range(2, ws.max_row+1):
            ws.cell(row=r, column=c).number_format = MONEY
    ws.freeze_panes = 'A2'
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---------------- Sheet 1: Summary ----------------
cash_amt = R.cash_total
traced_amt = R.traced_amt
missing_rows = [(d,r) for d,r in R.still_missing]
not_in_bank = [x for x in missing_rows if 'NOT FOUND' in x[1]]
summary_rows = [
    ('Period', '01-Apr-2025 to 31-Mar-2026', ''),
    ('Account', 'HDFC Bank 50200036753323 (Pashan branch)', ''),
    ('', '', ''),
    ('BANK POSITION', '', ''),
    ('Total credits (12 months)', R.tot_credits, ''),
    ('Total debits (12 months)', R.tot_debits, ''),
    ('Net surplus', R.tot_credits - R.tot_debits, ''),
    ('Closing balance 31-Mar-2026', 418719.12, ''),
    ('Balance continuity across months', 'OK - no breaks', ''),
    ('', '', ''),
    ('COLLECTIONS vs BANK', '', ''),
    ('Receipts in dues report', len(R.dues), sum(d['amount'] for d in R.dues)),
    ('Matched to bank by UTR/ref', len(R.matched_pairs), R.matched_due_amt),
    ('Traced by amount+date (VERIFY)', None, traced_amt),
    ('Genuinely unexplained receipts', len(missing_rows), sum(d['amount'] for d,_ in missing_rows)),
    ('  of which NOT in any bank statement', len(not_in_bank), sum(d['amount'] for d,_ in not_in_bank)),
    ('Cash receipts (expected)', None, cash_amt),
    ('Bank credits with no receipt (mostly FD cycles)', len(R.unmatched_bank_credits), sum(x['dp'] for x in R.unmatched_bank_credits)),
    ('', '', ''),
    ('EXPENSES vs BANK', '', ''),
    ('Vendor bills / net payable', len(R.exp), sum(e['net_pay'] for e in R.exp)),
    ('Settled per books', None, sum(e['settled'] for e in R.exp)),
    ('Settled AND matched to bank debits', None, R.matched_exp_amt),
    ('Settled but UNTRACED in bank', len(R.unmatched_exp), sum(e['settled'] for e in R.unmatched_exp)),
    ('Unpaid balance as of 25-Aug-2026', None, sum(e['bal'] for e in R.exp)),
    ('Bank debits with no vendor bill', len(R.deb_avail), sum(R.debits[i]['wd'] for i in R.deb_avail)),
    ('', '', ''),
    ('RED FLAGS', '', ''),
    ('Bounced inward cheque returns', 6, 26795.0),
    ('Manager salary paid off-book (suspected cash)', '~21 vouchers', 662841.78),
]
make_sheet(wb, 'Summary', ['Item', 'Count', 'Amount (Rs)'],
           summary_rows, [52, 16, 18], money_cols=(3,), tab_color='1F4E79')

# ---------------- Sheet 2: Receipt exceptions ----------------
def dt(s): return datetime.strptime(str(s)[:10], '%Y-%m-%d')

# Recompute the amount+date trace to split exceptions into "probable match" vs "missing"
work = R.unmatched_bank_credits[:]
traced_items = []          # (due, reason, bank_txn)
still_missing = list(R.still_missing)
missing_ids = {id(d) for d, _ in still_missing}
for d, r in R.unmatched_dues:
    if id(d) in missing_ids: continue
    hit = next((x for x in work if abs(x['dp'] - d['amount']) < 0.01
                and abs((x['date'] - dt(d['date'])).days) <= 15), None)
    if hit:
        work.remove(hit)
        traced_items.append((d, r, hit))

receipt_rows = []
for d, r, h in sorted(traced_items, key=lambda t: str(t[0]['receipt'])):
    receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                         d['amount'], 'TRACED BY AMOUNT+DATE (verify ref)',
                         f'{h["date"].strftime("%d/%m/%y")} Rs {h["dp"]:,.2f} {h["narr"][:55]}'))

remaining = [(d, r) for d, r in still_missing if 'cash' not in r.lower()]
used = set()
# combined-payment groups first (same UTR, sum match within Re 1)
by_ref = {}
for d, r in remaining:
    by_ref.setdefault(d['refno'], []).append(d)
consumed = set()
for ref, ds in by_ref.items():
    tot = sum(x['amount'] for x in ds)
    hit = next((x for x in work if abs(x['dp'] - tot) <= 1.0
                and abs((x['date'] - dt(ds[0]['date'])).days) <= 20), None)
    if hit is not None and len(ds) > 1:
        work.remove(hit); consumed.update(id(d) for d in ds)
        for d in ds:
            receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                                 d['amount'], 'COMBINED PAYMENT TRACED',
                                 f'Group total Rs {tot:,.2f} = bank {hit["date"].strftime("%d/%m/%y")} {hit["narr"][:60]}'))
for d, r in remaining:
    if id(d) in consumed: continue
    hits = [x for x in work if abs(x['dp'] - d['amount']) <= 1.0
            and abs((x['date'] - dt(d['date'])).days) <= 20]
    if len(hits) == 1:
        h = hits[0]; work.remove(h)
        cat = 'ROUNDING DIFF - TRACED' if abs(h['dp']-d['amount']) > 0.005 else 'TRACED (verify)'
        receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                             d['amount'], cat, f'{h["date"].strftime("%d/%m/%y")} {h["narr"][:70]}'))
    elif len(hits) > 1:
        receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                             d['amount'], 'AMBIGUOUS - verify manually',
                             f'{len(hits)} identical bank credits in window'))
    else:
        receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                             d['amount'], 'NOT IN ANY BANK STATEMENT', r))
# cash
for d, r in still_missing:
    if 'cash' in r.lower():
        receipt_rows.append((int(d['receipt']), str(d['date'])[:10], d['house'], d['payee'],
                             d['amount'], 'CASH RECEIPT (expected)', ''))

# ---------------- Sheet 2b: write receipt exceptions ----------------
make_sheet(wb, 'Unmatched Receipts',
           ['Receipt #', 'Receipt Date', 'Flat', 'Payee', 'Amount (Rs)', 'Exception Category', 'Detail / Bank Trace'],
           sorted(receipt_rows, key=lambda t: t[5]), [11, 13, 9, 26, 13, 34, 80], money_cols=(5,), tab_color='C00000')

# ---------------- Sheet 3: Untraced expense payments ----------------
pay_rows = []
for e in sorted(R.unmatched_exp, key=lambda e: -e['settled']):
    note = 'Suspected cash payment' if 'Gaikwad' in str(e['vendor']) else 'No matching bank debit found'
    pay_rows.append((int(e['vno']), str(e['date'])[:10], e['vendor'], e['ac'], e['settled'],
                     e['bal'], e['st_aug'], note))
make_sheet(wb, 'Untraced Payments',
           ['Voucher #', 'Bill Date', 'Vendor', 'Expense A/C', 'Settled (Rs)', 'Balance (Rs)', 'Status 25-Aug-26', 'Note'],
           pay_rows, [10, 12, 38, 30, 13, 13, 16, 30], money_cols=(5, 6), tab_color='C00000')

# ---------------- Sheet 4: Bounced cheques ----------------
bounce_rows = []
for i in R.deb_avail:
    x = R.debits[i]
    n = x['narr'].upper()
    if 'CHQ DEP RET' in n or 'I/W CHQ RET' in n or ('CHQ RET' in n):
        bounce_rows.append((x['date'].strftime('%d/%m/%Y'), x['wd'], x['narr'], x['month']))
make_sheet(wb, 'Bounced Cheques',
           ['Bank Date', 'Amount (Rs)', 'Narration', 'Statement Month'],
           bounce_rows, [14, 14, 80, 16], money_cols=(2,), tab_color='ED7D31')

# ---------------- Sheet 5: Bank debits with no vendor bill ----------------
deb_rows = []
for i in sorted(R.deb_avail, key=lambda i: R.debits[i]['date']):
    x = R.debits[i]
    deb_rows.append((x['date'].strftime('%d/%m/%Y'), x['wd'], R.classify_debit(x), x['narr']))
make_sheet(wb, 'Bank Debits w-o Bill',
           ['Bank Date', 'Amount (Rs)', 'Classification', 'Narration'],
           deb_rows, [14, 14, 36, 85], money_cols=(2,), tab_color='ED7D31')

# ---------------- Sheet 6: Bank credits with no receipt ----------------
cred_rows = []
for x in sorted(R.unmatched_bank_credits, key=lambda x: x['date']):
    cred_rows.append((x['date'].strftime('%d/%m/%Y'), x['dp'], R.classify_credit(x), x['narr']))
make_sheet(wb, 'Bank Credits w-o Receipt',
           ['Bank Date', 'Amount (Rs)', 'Classification', 'Narration'],
           cred_rows, [14, 14, 44, 85], money_cols=(2,), tab_color='70AD47')

out = os.path.join(BASE, 'Reconciliation_Exceptions_FY2025-26.xlsx')
wb.save(out)
print('Saved:', out)
print('Sheets:', wb.sheetnames)
print('Receipt exception rows :', len(receipt_rows))
print('Untraced payment rows  :', len(pay_rows))
print('Bounced cheque rows    :', len(bounce_rows))
print('Debit w/o bill rows    :', len(deb_rows))
print('Credit w/o receipt rows:', len(cred_rows))

